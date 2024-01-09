from flask import Flask, render_template
from datetime import datetime
from openpyxl import load_workbook

app = Flask(__name__)


@app.route('/')
def home_page():
    current_year = datetime.now().year
    age = current_year - 1996
    return render_template('index.html', age=age)


@app.route('/<project_title>')
def portfolio_details(project_title):
    """GET RESULTS WHERE THE SELECTED PROJECT TITLE IS EQUAL TO THE PROJECT TITLE IN THE EXCEL"""
    wb = load_workbook('./instance/new-projects.xlsx')
    ws = wb.active
    for row in ws.iter_rows():
        if row[1].value == project_title:
            selected_table_row = row
            break
    else:
        selected_table_row = ()

    return render_template('portfolio-details.html', selected_project=selected_table_row)


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
